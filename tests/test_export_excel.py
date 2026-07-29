import os
from openpyxl import load_workbook

from models.student import Student
from services.report_service import export_excel


def test_export_excel(tmp_path):
    students = [
        Student("SV001", "An", 20, "CNTT", 9, 8, 7),
        Student("SV002", "Bình", 21, "CNTT", 7, 7, 7)
    ]

    file_path = tmp_path / "students.xlsx"

    export_excel(students, str(file_path))

    assert os.path.exists(file_path)

    wb = load_workbook(file_path)
    ws = wb.active

    # Header + 2 sinh viên
    assert ws.max_row == 3

    assert ws["A2"].value == "SV001"
    assert ws["B2"].value == "An"

    assert ws["A3"].value == "SV002"
    assert ws["B3"].value == "Bình"
